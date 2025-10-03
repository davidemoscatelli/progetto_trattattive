import json
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.db.models import Q
from django.contrib.auth.models import User
from .models import Trattativa, Task, Commento
from .forms import TrattativaForm, CommentoForm, TaskForm

@login_required
def homepage(request):
    tasks_urgenti = Task.objects.filter(
        assegnato_a=request.user, 
        priorita=3,
        completato=False
    )
    trattative_in_corso_count = Trattativa.objects.exclude(stato__in=['Vinta', 'Persa']).count()
    trattative_vinte_count = Trattativa.objects.filter(stato='Vinta').count()
    valore_in_corso = Trattativa.objects.exclude(stato__in=['Vinta', 'Persa']).aggregate(Sum('valore'))['valore__sum'] or 0
    context = {
        'tasks_urgenti': tasks_urgenti,
        'trattative_in_corso_count': trattative_in_corso_count,
        'trattative_vinte_count': trattative_vinte_count,
        'valore_in_corso': valore_in_corso,
    }
    return render(request, 'trattative/homepage.html', context)

@login_required
def kanban_board(request):
    trattative = Trattativa.objects.all()
    kanban_data = {
        stato[0]: trattative.filter(stato=stato[0]) 
        for stato in Trattativa.STATO_CHOICES
    }
    context = {'kanban_data': kanban_data}
    return render(request, 'trattative/kanban_board.html', context)

@login_required
def kpi_dashboard(request):
    trattative_vinte = Trattativa.objects.filter(stato='Vinta')
    trattative_perse = Trattativa.objects.filter(stato='Persa')
    totali_concluse = trattative_vinte.count() + trattative_perse.count()
    tasso_conversione = (trattative_vinte.count() / totali_concluse * 100) if totali_concluse > 0 else 0
    valore_totale_vinto = trattative_vinte.aggregate(Sum('valore'))['valore__sum'] or 0
    context = {
        'trattative_in_corso': Trattativa.objects.exclude(stato__in=['Vinta', 'Persa']).count(),
        'valore_in_corso': Trattativa.objects.exclude(stato__in=['Vinta', 'Persa']).aggregate(Sum('valore'))['valore__sum'] or 0,
        'tasso_conversione': tasso_conversione,
        'valore_totale_vinto': valore_totale_vinto,
    }
    return render(request, 'trattative/kpi_dashboard.html', context)

@login_required
def trattativa_detail(request, pk):
    trattativa = get_object_or_404(Trattativa, pk=pk)
    commenti = trattativa.commenti.all()
    if request.method == 'POST':
        form = CommentoForm(request.POST)
        if form.is_valid():
            nuovo_commento = form.save(commit=False)
            nuovo_commento.trattativa = trattativa
            nuovo_commento.autore = request.user
            nuovo_commento.save()
            return redirect('trattativa_detail', pk=trattativa.pk)
    else:
        form = CommentoForm()
    context = {'trattativa': trattativa, 'commenti': commenti, 'form': form}
    return render(request, 'trattative/trattativa_detail.html', context)

@login_required
def trattativa_create(request):
    if request.method == 'POST':
        form = TrattativaForm(request.POST)
        if form.is_valid():
            # Creiamo l'oggetto in memoria senza salvarlo subito
            trattativa = form.save(commit=False)
            
            # --- LOGICA DI ASSEGNAZIONE AUTOMATICA ---
            # Se l'utente non ha selezionato un responsabile, lo assegniamo a lui
            if not trattativa.responsabile:
                trattativa.responsabile = request.user
            
            # Ora salviamo l'oggetto completo nel database
            trattativa.save()
            
            return redirect('kanban_board')
    else:
        form = TrattativaForm()
    return render(request, 'trattative/trattativa_form.html', {'form': form, 'tipo_azione': 'Nuova'})

@login_required
def trattativa_update(request, pk):
    trattativa = get_object_or_404(Trattativa, pk=pk)
    if request.method == 'POST':
        form = TrattativaForm(request.POST, instance=trattativa)
        if form.is_valid():
            form.save()
            return redirect('kanban_board')
    else:
        form = TrattativaForm(instance=trattativa)
    return render(request, 'trattative/trattativa_form.html', {'form': form, 'tipo_azione': 'Modifica'})

@login_required
def trattativa_delete(request, pk):
    trattativa = get_object_or_404(Trattativa, pk=pk)
    if request.method == 'POST':
        trattativa.delete()
        return redirect('kanban_board')
    return render(request, 'trattative/trattativa_confirm_delete.html', {'trattativa': trattativa})

@login_required
def task_list(request):
    tasks = Task.objects.filter(completato=False)
    context = {'tasks': tasks}
    return render(request, 'trattative/task_list.html', context)

@login_required
def task_create(request):
    if request.method == 'POST':
        form = TaskForm(request.user, request.POST)
        if form.is_valid():
            task = form.save(commit=False)
            task.creato_da = request.user
            task.save()
            return redirect('task_list')
    else:
        form = TaskForm(request.user)
    return render(request, 'trattative/task_form.html', {'form': form})

@login_required
def task_update(request, pk):
    task = get_object_or_404(Task, pk=pk)
    if request.method == 'POST':
        form = TaskForm(request.user, request.POST, instance=task)
        if form.is_valid():
            form.save()
            return redirect('task_list')
    else:
        form = TaskForm(request.user, instance=task)
    return render(request, 'trattative/task_form.html', {'form': form})

@login_required
def task_complete(request, pk):
    task = get_object_or_404(Task, pk=pk)
    if request.method == 'POST':
        task.completato = True
        task.save()
    return redirect('task_list')

@login_required
def profilo(request):
    return render(request, 'trattative/profilo.html')

@login_required
def esporta_trattative_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Elenco Trattative"

    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center')
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    colonne = ['Titolo', 'Cliente', 'Valore', 'Stato', 'Responsabile', 'Data Creazione']
    ws.append(colonne)
    for col_num, column_title in enumerate(colonne, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill

    trattative = Trattativa.objects.all()
    for trattativa in trattative:
        ws.append([
            trattativa.titolo,
            trattativa.cliente,
            trattativa.valore,
            trattativa.stato,
            trattativa.responsabile.username if trattativa.responsabile else 'N/A',
            trattativa.data_creazione.strftime("%d/%m/%Y")
        ])
        # Applica il formato valuta alla cella del valore (Colonna C)
        ws[f'C{ws.max_row}'].number_format = '€ #,##0.00'

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="elenco_trattative.xlsx"'
    wb.save(response)
    return response

# --- VISTE API ---
@csrf_exempt
@login_required
def update_trattativa_status(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        trattativa_id = data.get('trattativa_id')
        nuovo_stato = data.get('nuovo_stato')
        try:
            trattativa = Trattativa.objects.get(pk=trattativa_id)
            trattativa.stato = nuovo_stato
            trattativa.save()
            return JsonResponse({'status': 'success', 'message': 'Stato aggiornato'})
        except Trattativa.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Trattativa non trovata'}, status=404)
    return JsonResponse({'status': 'error', 'message': 'Metodo non valido'}, status=400)

@login_required
def chart_data_api(request):
    vinte_count = Trattativa.objects.filter(stato='Vinta').count()
    perse_count = Trattativa.objects.filter(stato='Persa').count()
    data = {
        'labels': ['Vinte', 'Perse'],
        'data': [vinte_count, perse_count],
    }
    return JsonResponse(data)

@login_required
def trattativa_list(request):
    queryset = Trattativa.objects.all().order_by('-ultima_modifica')

    # Recupera tutti i parametri di filtro
    query = request.GET.get('q')
    stato_filter = request.GET.get('stato')
    responsabile_filter = request.GET.get('responsabile')
    collaboratore_filter = request.GET.get('collaboratore')
    
    # Filtro di ricerca per testo
    if query:
        queryset = queryset.filter(
            Q(titolo__icontains=query) |
            Q(cliente__icontains=query)
        ).distinct()

    # Filtro per stato
    if stato_filter:
        queryset = queryset.filter(stato=stato_filter)
        
    # Filtro per responsabile
    if responsabile_filter:
        queryset = queryset.filter(responsabile_id=responsabile_filter)
        
    # Filtro per collaboratore
    if collaboratore_filter:
        queryset = queryset.filter(collaboratori__id=collaboratore_filter)
        
    context = {
        'trattative': queryset,
        'stati_disponibili': Trattativa.STATO_CHOICES,
        'utenti_disponibili': User.objects.all(), # Passa tutti gli utenti al template
    }
    return render(request, 'trattative/trattativa_list.html', context)


@login_required
def kpi_dashboard(request):
    # Riepilogo trattative (già esistente)
    trattative_vinte = Trattativa.objects.filter(stato='Vinta')
    valore_totale_vinto = trattative_vinte.aggregate(Sum('valore'))['valore__sum'] or 0
    trattative_in_corso_count = Trattativa.objects.exclude(stato__in=['Vinta', 'Persa']).count()
    
    # --- NUOVA LOGICA PER LA PIPELINE PESATA ---

    # Valore al 25% (Da Contattare, Contattato)
    valore_25 = Trattativa.objects.filter(
        stato__in=['Da Contattare', 'Contattato']
    ).aggregate(Sum('valore'))['valore__sum'] or 0

    # Valore al 50% (Demo, Preventivo Tecnico)
    valore_50 = Trattativa.objects.filter(
        stato__in=['Demo', 'Preventivo Tecnico']
    ).aggregate(Sum('valore'))['valore__sum'] or 0

    # Valore al 75% (Preventivo Commerciale)
    valore_75 = Trattativa.objects.filter(
        stato='Preventivo Commerciale'
    ).aggregate(Sum('valore'))['valore__sum'] or 0
    
    # Calcolo del valore totale pesato della pipeline
    valore_pesato = (valore_25 * 0.25) + (valore_50 * 0.50) + (valore_75 * 0.75)

    context = {
        'trattative_in_corso_count': trattative_in_corso_count,
        'valore_totale_vinto': valore_totale_vinto,
        
        # Nuovi valori per il template
        'valore_25': valore_25,
        'valore_50': valore_50,
        'valore_75': valore_75,
        'valore_pesato': valore_pesato,
    }
    return render(request, 'trattative/kpi_dashboard.html', context)